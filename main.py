from concurrent.futures import thread
import os, openpyxl
from pickletools import optimize
import pandas as pd
from openpyxl.utils.cell import get_column_letter
from PIL import Image as Image2
import threading

#This python tool needs the pip install of some of the libraries above in order to function.

fichExcel = "InventaireDesTissus.xlsx" #Your excel filename
hauteur = 200 #Height of the picture in excel.
largeur = 200 #Widhth of the picture in excel.
picturesPath = "./photos/" #Here you put the path where your images are stored.

#Declaration of lists used by the code
sheets = []
photos = []
nomsPhotos = []
threads = []


def nomFichiers():
    #This functions looks for the pictures to add to excel
    files = os.listdir(path=picturesPath) 
    for file in files :
        nomsPhotos.append(file) #One first list to store the path of each pictures.
        photos.append(file.strip("DSC0").strip(".JPG")) #In my case, appends a second list stripping some stuff off the name.

def findAndWrite(sheet):
    #This functions matches the number in excel with the picture number and inserts the picture in the corresponding cell.
    wb = openpyxl.load_workbook(fichExcel)
    s = wb[sheet]
    r = s.max_row
    s.column_dimensions['A'].width = 35
    print('Importation des images sur le classeur "'+sheet+'".') #Importation of images in excel sheet

    #Keep in mind, the loop below only loops the A1 column for the lenght of the rows.
    for k in range(r) :
        cell1 = str(s[get_column_letter(1)+str(k+1)].value)
        for num in photos :
            if str(cell1) == num :
                s.row_dimensions[k+1].height = hauteur-20
                nom = resizeImage(num)
                img = openpyxl.drawing.image.Image(nom)
                s.add_image(img, 'A'+str(k+1))
    print('"'+sheet+'"'+' : terminé.')
    wb.save(fichExcel)

def excel():
    #Function below uses threading to use a thread for each sheet of the excel file.
    print('Importation de '+str(len(photos))+' images dans '+fichExcel+'. Ceci peux prendre un certain temps.')
    print("")
    print("")
    df = pd.ExcelFile(fichExcel)
    sheets = df.sheet_names
    for sheet in sheets :
        t = threading.Thread(target=findAndWrite(sheet))
        threads.append(t)
        t.start()
    for ts in threads :
        ts.join()
                    

                    
def resizeImage(number) :
    #This one resizes every pictures in your pictures folder. #Warning, this changes the source images, make sure to have a backup.
    nom = 'DSC0'+number+'.JPG'
    img = Image2.open(picturesPath+nom)
    img = img.resize((int(hauteur),int(largeur)), Image2.Dither.NONE)
    img.save(picturesPath+nom, optimize=True, quality=100)
    return picturesPath+nom


#main
nomFichiers()
excel()